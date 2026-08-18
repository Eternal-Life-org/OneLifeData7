#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
translate.py - 用 翻译表.xlsx 覆盖 objects/*.txt 的名称行(第2行)

用法:
    cd OneLifeData7
    python translate.py              # 同步工具菜单: 新增物品/导入修改/应用翻译/撤回
    python translate.py --mode 1     # 直接执行: 1=翻译英文 2=翻译中文 3=追加英文(中文+英文)
    python translate.py --mode 3 --dry-run    # 只扫描不覆盖(测试用, 附带改动统计)
    python translate.py --mode 3 --ignore-errors  # 输出异常但仍执行覆盖

菜单四个按钮(每次写入前自动备份, 可撤回一次):
    1. 新增物品 (本地 -> 翻译表): objects/ 里表上没有的编号, 追加为新行(B英文/C中文/D后缀)
    2. 导入修改 (本地 -> 翻译表): 本地第2行与表(实装B/C/D)不一致的, 用本地覆盖 B/C/D
    3. 应用翻译 (翻译表 -> 本地): mode 3 追加英文(中文+英文)
    4. 撤回: 回退到上一次操作前的状态
    备份在 OneLifeData7/sync_backup/

xlsx 结构 (Elife sheet, 按表头识别列):
    A=key(物品id)  B=English  C=简体中文（实装）  D=后缀（实装）
    E=简体中文（录入区）  F=后缀（录入区）  —— 录入区仅供调整, 本脚本不读取
    G=修改状态(公式, 只比名字)  H=备注

只有"实装"两列会被录入:
    C = 中文名(不含后缀)    D = 后缀(以#开头, 可为空)

后缀规范: 每个 # 后必须有一个空格(行业规范)。脚本读表时自动规范化,
'#Flowering' 与 '# Flowering' 视为同一后缀, 输出统一为 '# Flowering'。

三种模式构建的第2行:
    1 翻译英文:  English + 后缀
    2 翻译中文:  Chinese + 后缀
    3 追加英文:  Chinese + English + 后缀   (名字与后缀拼接均无额外空格)

扫描规则(覆盖前) —— 只检查名字, 不检查后缀:
    - object 文件不存在       -> warning (跳过, 不阻止覆盖)
    - xlsx 缺少该模式的翻译列 -> warning (跳过该行不覆盖)
    - object 第2行为空        -> 异常
    - xlsx 中英文均为空       -> 异常
    有异常时默认退出不覆盖; 加 --ignore-errors 则输出异常后继续覆盖
    (用 xlsx 数据覆盖 object, 可修复空行)。

兼容旧格式表(表头 key/English/Chinese/Label 或无表头的 B/C/D 固定列)。
"""

import sys
import os
import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
import openpyxl

NS  = '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}'
RNS = '{http://schemas.openxmlformats.org/officeDocument/2006/relationships}'

# Elife sheet 名称
SHEET_NAME = 'Elife'
# 默认文件/目录 (相对运行时 cwd)
DEFAULT_XLSX = '翻译表.xlsx'
DEFAULT_OBJ_DIR = 'objects'


# ---------- xlsx 读取 (不依赖 openpyxl, 直接解析 XML) ----------

def _load_shared(z):
    """sharedStrings 可选: openpyxl 等工具保存的表用内联字符串, 没有该文件"""
    if 'xl/sharedStrings.xml' not in z.namelist():
        return []
    ss_tree = ET.fromstring(z.read('xl/sharedStrings.xml'))
    return [''.join(t.text or '' for t in si.iter(f'{NS}t'))
            for si in ss_tree.findall(f'{NS}si')]


def _find_elife_sheet_path(z):
    """在 xlsx 里找 Elife sheet 对应的 worksheet xml 路径"""
    wb = ET.fromstring(z.read('xl/workbook.xml'))
    rid_target = {}
    rels = ET.fromstring(z.read('xl/_rels/workbook.xml.rels'))
    for rel in rels:
        rid_target[rel.get('Id')] = rel.get('Target')

    for sh in wb.findall(f'{NS}sheets/{NS}sheet'):
        if sh.get('name') == SHEET_NAME:
            rid = sh.get(f'{RNS}id')
            target = rid_target.get(rid)
            if target:
                if not target.startswith('/'):
                    target = 'xl/' + target
                else:
                    target = target.lstrip('/')
                return target
    return None


def _cell_value(c, shared):
    """取一个 <c> 单元格的值(支持共享字符串/内联字符串/数字)"""
    t = c.get('t')
    v = c.find(f'{NS}v')
    isn = c.find(f'{NS}is')
    if t == 's' and v is not None:
        return shared[int(v.text)]
    if isn is not None:
        return ''.join(tt.text or '' for tt in isn.iter(f'{NS}t'))
    if v is not None:
        return v.text or ''
    return ''


def _map_columns(header_cells):
    """按表头行识别列; 返回 (key, english, chinese, label) 列字母, 识别失败返回 None"""
    def find(pred):
        for col, v in header_cells.items():
            if pred(v.strip()):
                return col
        return None

    eng = find(lambda v: v.lower() == 'english')
    chn = find(lambda v: ('实装' in v and '中文' in v) or v.lower() == 'chinese')
    lab = find(lambda v: ('实装' in v and '后缀' in v) or v.lower() == 'label')
    key = find(lambda v: v.lower() == 'key')
    if eng and chn:
        return (key or 'A', eng, chn, lab)
    return None


def read_xlsx(xlsx_path):
    """读取 Elife sheet, 返回 dict[key] = {english, chinese, label}
    列按表头识别(实装列); 无表头或表头不认识时回退旧格式固定列 A/B/C/D"""
    z = zipfile.ZipFile(xlsx_path)
    shared = _load_shared(z)

    sheet_path = _find_elife_sheet_path(z)
    if sheet_path is None:
        raise RuntimeError(f"xlsx 里找不到名为 '{SHEET_NAME}' 的 sheet")

    sh_tree = ET.fromstring(z.read(sheet_path))
    xml_rows = sh_tree.findall(f'{NS}sheetData/{NS}row')

    translations = {}
    cols = None
    for row in xml_rows:
        cells = {}
        for c in row.findall(f'{NS}c'):
            ref = c.get('r')
            col = re.match(r'([A-Z]+)', ref).group(1)
            cells[col] = _cell_value(c, shared)

        if cols is None:
            # 首行: 是表头则识别列, 是数据(key为纯数字)则按旧格式固定列
            if re.match(r'^\d+$', (cells.get('A') or '').strip()):
                cols = ('A', 'B', 'C', 'D')
            else:
                cols = _map_columns(cells) or ('A', 'B', 'C', 'D')
                continue

        key = clean_cell(cells.get(cols[0], ''))
        # 只接受正整数 key (object id)
        if not re.match(r'^\d+$', key):
            continue
        translations[key] = {
            'english': clean_cell(cells.get(cols[1], '')),
            'chinese': clean_cell(cells.get(cols[2], '')),
            'label':   normalize_suffix(clean_cell(cells.get(cols[3], ''))) if cols[3] else '',
        }
    return translations


def clean_cell(s):
    """单元格清洗: 去内部回车换行, 去前后空白"""
    if s is None:
        return ''
    s = str(s)
    s = s.replace('\r', '').replace('\n', '')
    return s.strip()


def normalize_suffix(s):
    """后缀规范: 每个 # 后必须有一个空格(行业规范, 提升码风)
    '#Flowering' -> '# Flowering';  '#W #cart' -> '# W # cart'
    行尾的 # 不处理(后面没内容)"""
    return re.sub(r'#(?=\S)', '# ', s or '')


# ---------- object 文件读写 ----------

def read_obj_line2(obj_path):
    """读 object 第2行(已去 \\r\\n); 文件不存在或不足2行返回 None"""
    try:
        with open(obj_path, encoding='utf-8') as f:
            content = f.read()
    except FileNotFoundError:
        return None
    content = content.replace('\r\n', '\n').replace('\r', '\n')
    lines = content.split('\n')
    if len(lines) < 2:
        return None
    line2 = lines[1]
    if line2.strip() == '':
        return ''
    return line2


def write_obj_line2(obj_path, new_line2):
    """覆盖 object 第2行, 保留其余行, 统一 LF"""
    with open(obj_path, encoding='utf-8') as f:
        content = f.read()
    content = content.replace('\r\n', '\n').replace('\r', '\n')
    lines = content.split('\n')
    if len(lines) < 2:
        return False
    lines[1] = new_line2
    with open(obj_path, 'w', encoding='utf-8', newline='') as f:
        f.write('\n'.join(lines))
    return True


# ---------- 名称构建 ----------

def mode_label(mode):
    return {1: '翻译英文', 2: '翻译中文', 3: '追加英文(中文+英文)'}[mode]


def mode_needs(mode):
    """该模式需要的翻译列 (用于扫描提示缺失)"""
    if mode == 1:
        return ['english']
    if mode == 2:
        return ['chinese']
    if mode == 3:
        return ['chinese', 'english']
    return []


def has_name(mode, t):
    """该模式下是否有可用名称 (决定是否跳过覆盖)
    mode 3 缺一列时用另一列 fallback, 都缺才跳过"""
    if mode == 1:
        return bool(t['english'])
    if mode == 2:
        return bool(t['chinese'])
    if mode == 3:
        return bool(t['chinese']) or bool(t['english'])
    return False


def build_name(mode, t):
    """构建名称部分(不含后缀)"""
    if mode == 1:
        return t['english']
    if mode == 2:
        return t['chinese']
    if mode == 3:
        cn = t['chinese']
        en = t['english']
        if cn and en:
            # 中英文都有: 一致则保留一份, 否则拼接
            return cn if cn == en else cn + en
        # 只有其一: 用有的那个 (mode 3 对单语物品的 fallback)
        return cn or en
    return ''


def build_line2(mode, t):
    """构建完整第2行: 名字 + 后缀 (后缀自带 #, 无空格)"""
    return build_name(mode, t) + t['label']


# ---------- 扫描 ----------

def scan(translations, mode, obj_dir):
    """扫描所有 key, 返回 (errors, warnings) —— 只检查名字, 不检查后缀"""
    errors = []
    warnings = []
    needed = mode_needs(mode)

    for key in sorted(translations.keys(), key=lambda x: int(x)):
        t = translations[key]
        obj_path = os.path.join(obj_dir, f'{key}.txt')
        line2 = read_obj_line2(obj_path)

        # 1. object 文件不存在/行数不足 -> warning (xlsx 可有多余条目, 跳过即可)
        if line2 is None:
            warnings.append(f"key {key}: object 文件不存在或行数不足2行, 跳过")
            continue
        # object 第2行为空 -> 异常
        if line2 == '':
            errors.append(f"key {key}: object 第2行为空")
            continue

        # xlsx 中英文均为空 -> 异常 (无翻译数据, 不覆盖)
        if not t['english'] and not t['chinese']:
            errors.append(f"key {key}: xlsx 中英文均为空")
            continue

        # 2. xlsx 缺翻译列 -> warning (只缺一列, mode 3 可 fallback)
        missing = [c for c in needed if not t[c]]
        if missing:
            warnings.append(
                f"key {key}: xlsx 缺少 {','.join(missing)} 翻译")

    return errors, warnings


# ---------- 覆盖 ----------

def translate(translations, mode, obj_dir):
    """执行覆盖, 返回成功覆盖的行数"""
    count = 0
    skipped = 0
    for key in sorted(translations.keys(), key=lambda x: int(x)):
        t = translations[key]
        # 无可用名称才跳过 (mode 3 缺一列时用另一列 fallback)
        if not has_name(mode, t):
            skipped += 1
            continue
        obj_path = os.path.join(obj_dir, f'{key}.txt')
        if not os.path.isfile(obj_path):
            skipped += 1
            continue
        if write_obj_line2(obj_path, build_line2(mode, t)):
            count += 1
        else:
            skipped += 1
    return count, skipped


# ---------- 同步工具: 备份 / 撤回 ----------

BACKUP_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'sync_backup')
BACKUP_XLSX = os.path.join(BACKUP_DIR, 'table_backup.xlsx')
BACKUP_OBJ = os.path.join(BACKUP_DIR, 'objects')
BACKUP_MARK = os.path.join(BACKUP_DIR, 'last_action.txt')


def save_table_backup():
    """备份翻译表, 标记最近一次动作为 table"""
    os.makedirs(BACKUP_DIR, exist_ok=True)
    shutil.copy2(DEFAULT_XLSX, BACKUP_XLSX)
    with open(BACKUP_MARK, 'w', encoding='utf-8') as f:
        f.write('table')


def save_objects_backup(keys):
    """备份将被改动的 object 文件, 标记最近一次动作为 objects"""
    os.makedirs(BACKUP_OBJ, exist_ok=True)
    for fn in os.listdir(BACKUP_OBJ):
        os.remove(os.path.join(BACKUP_OBJ, fn))
    for k in keys:
        shutil.copy2(os.path.join(DEFAULT_OBJ_DIR, f'{k}.txt'),
                     os.path.join(BACKUP_OBJ, f'{k}.txt'))
    with open(BACKUP_MARK, 'w', encoding='utf-8') as f:
        f.write('objects')


def action_undo():
    """撤回: 回退到上一次操作前的状态"""
    if not os.path.isfile(BACKUP_MARK):
        print('[撤回] 没有可撤回的操作')
        return
    with open(BACKUP_MARK, encoding='utf-8') as f:
        last = f.read().strip()
    if last == 'table':
        if not os.path.isfile(BACKUP_XLSX):
            print('[撤回] 表备份缺失, 无法撤回')
            return
        shutil.copy2(BACKUP_XLSX, DEFAULT_XLSX)
        os.remove(BACKUP_MARK)
        print('[撤回] 翻译表已回退到上一次操作前')
    elif last == 'objects':
        files = [f for f in os.listdir(BACKUP_OBJ) if f.endswith('.txt')]
        if not files:
            print('[撤回] 对象备份为空, 无法撤回')
            return
        for f in files:
            shutil.copy2(os.path.join(BACKUP_OBJ, f),
                         os.path.join(DEFAULT_OBJ_DIR, f))
        os.remove(BACKUP_MARK)
        print(f'[撤回] {len(files)} 个 object 已回退到上一次操作前')
    else:
        print(f'[撤回] 未知备份类型: {last}')


# ---------- 同步工具: 新增 / 导入修改 ----------

def split_cn_en(name):
    """'轨道终点Track End' -> ('轨道终点', 'Track End'); 纯中文/纯英文也能处理"""
    name = name.strip()
    i = 0
    while i < len(name) and ord(name[i]) > 127:
        i += 1
    return name[:i].rstrip(), name[i:].lstrip()


def parse_line2(line2):
    """'轨道终点Track End# W cart' -> (cn, en, suffix); 无#时后缀为空"""
    idx = line2.find('#')
    if idx < 0:
        name, suf = line2, ''
    else:
        name, suf = line2[:idx], line2[idx:]
    cn, en = split_cn_en(name)
    return cn, en, suf


def load_table():
    """加载翻译表(读写), 返回 (wb, ws, {key: 行号})"""
    wb = openpyxl.load_workbook(DEFAULT_XLSX)
    ws = wb[SHEET_NAME]
    key2row = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        k = str(row[0].value or '').strip()
        if k.isdigit():
            key2row[k] = row[0].row
    return wb, ws, key2row


def action_add():
    """1. 新增物品: objects/ 里表上没有的编号, 追加为新行(B英文/C中文/D后缀)"""
    wb, ws, key2row = load_table()
    added, skipped = [], []
    for fn in sorted(os.listdir(DEFAULT_OBJ_DIR)):
        if not fn.endswith('.txt'):
            continue
        key = fn[:-4]
        if not key.isdigit() or key in key2row:
            continue
        line2 = read_obj_line2(os.path.join(DEFAULT_OBJ_DIR, fn))
        if line2 is None or line2 == '':
            skipped.append((key, '第2行缺失或为空'))
            continue
        cn, en, suf = parse_line2(line2)
        if not cn and not en:
            skipped.append((key, '名字为空'))
            continue
        r = ws.max_row + 1
        ws.cell(row=r, column=1).value = int(key)
        ws.cell(row=r, column=2).value = en or None
        ws.cell(row=r, column=3).value = cn or None
        ws.cell(row=r, column=4).value = suf or None
        ws.cell(row=r, column=5).value = cn or None
        ws.cell(row=r, column=6).value = suf or None
        ws.cell(row=r, column=7).value = '=IF(AND(C{r}=E{r},D{r}=F{r}),"未修改","有修改")'.format(r=r)
        added.append(key)
    if added:
        save_table_backup()
        wb.save(DEFAULT_XLSX)
    print(f'[新增] 加入 {len(added)} 行: {", ".join(added[:20])}{" ..." if len(added) > 20 else ""}')
    if skipped:
        print(f'[新增] 跳过 {len(skipped)} 个: {skipped[:10]}{" ..." if len(skipped) > 10 else ""}')


def action_import():
    """2. 导入修改: 本地第2行与表(实装B/C/D)不一致的, 用本地覆盖 B/C/D"""
    wb, ws, key2row = load_table()
    changed = []
    for k in sorted(key2row, key=int):
        r = key2row[k]
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        d = ws.cell(row=r, column=4).value
        t = {
            'english': str(b).strip() if b is not None else '',
            'chinese': str(c).strip() if c is not None else '',
            'label': normalize_suffix(str(d).strip() if d is not None else ''),
        }
        line2 = read_obj_line2(os.path.join(DEFAULT_OBJ_DIR, f'{k}.txt'))
        if line2 is None:
            continue  # object 不存在, 不管
        if build_line2(3, t) == line2:
            continue  # 表与本地一致
        cn, en, suf = parse_line2(line2)
        ws.cell(row=r, column=2).value = en or None
        ws.cell(row=r, column=3).value = cn or None
        ws.cell(row=r, column=4).value = suf or None
        changed.append(k)
    if changed:
        save_table_backup()
        wb.save(DEFAULT_XLSX)
    print(f'[导入修改] 用本地覆盖 {len(changed)} 行: {", ".join(changed[:20])}{" ..." if len(changed) > 20 else ""}')


# ---------- 执行翻译 ----------

def run_mode(mode, ignore_errors=False, backup_objects=False, dry_run=False):
    """执行翻译(1/2/3); backup_objects=True 时先把将改动的 object 备份(供撤回)"""
    xlsx_path = DEFAULT_XLSX
    obj_dir = DEFAULT_OBJ_DIR

    print(f"模式: {mode_label(mode)}")
    print(f"读取 xlsx: {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        print(f"错误: 找不到 xlsx 文件 '{xlsx_path}'")
        return

    print("解析 xlsx ...")
    translations = read_xlsx(xlsx_path)
    print(f"xlsx 共 {len(translations)} 条 key")

    print("扫描 object 第2行 ...")
    errors, warnings = scan(translations, mode, obj_dir)

    if warnings:
        print(f"\n--- warnings: {len(warnings)} 条 ---")
        for w in warnings:
            print(f"  [WARN] {w}")

    if errors:
        print(f"\n--- 异常: {len(errors)} 条 ---")
        for e in errors:
            print(f"  [ERROR] {e}")
        if ignore_errors:
            print(f"\n--ignore-errors: 忽略 {len(errors)} 处异常, 继续覆盖 "
                  f"(将用 xlsx 数据覆盖 object, 可修复空行)。")
        else:
            print(f"\n扫描发现 {len(errors)} 处异常, 终止覆盖。"
                  f"请先修复, 或加 --ignore-errors 忽略异常继续覆盖。")
            return

    print(f"\n扫描通过: {len(translations)} 条, warning {len(warnings)} 条, "
          f"异常 {len(errors)} 条"
          f"{' (已忽略)' if ignore_errors else ''}。")

    changed = []
    for key in sorted(translations.keys(), key=lambda x: int(x)):
        t = translations[key]
        if not has_name(mode, t):
            continue
        line2 = read_obj_line2(os.path.join(obj_dir, f'{key}.txt'))
        if line2 is None or line2 == '':
            continue
        if line2 != build_line2(mode, t):
            changed.append(key)
    print(f"\n预览: 覆盖后将改动 {len(changed)} 个 object 的第2行"
          f"{' (--dry-run, 未写入)' if dry_run else ''}")
    for key in changed[:10]:
        t = translations[key]
        line2 = read_obj_line2(os.path.join(obj_dir, f'{key}.txt'))
        print(f"  {key}: {line2!r} -> {build_line2(mode, t)!r}")

    if dry_run:
        print("--dry-run 模式, 不执行覆盖。")
        return

    if backup_objects and changed:
        save_objects_backup(changed)

    print("开始覆盖 ...")
    count, skipped = translate(translations, mode, obj_dir)
    print(f"完成: 覆盖 {count} 个 object, 跳过 {skipped} 个。")


# ---------- 同步工具菜单 ----------

def sync_menu():
    """无参数启动时的菜单: 新增物品/导入修改/应用翻译/撤回"""
    wb = openpyxl.load_workbook(DEFAULT_XLSX, read_only=True)
    ws = wb[SHEET_NAME]
    n_obj = len([f for f in os.listdir(DEFAULT_OBJ_DIR)
                 if f.endswith('.txt') and f[:-4].isdigit()])
    print(f'翻译表: {ws.max_row - 1} 行 | 本地 objects: {n_obj} 个')
    wb.close()

    while True:
        print()
        print('======== 翻译表同步工具 ========')
        print('1. 新增物品   (本地 -> 翻译表)')
        print('2. 导入修改   (本地 -> 翻译表)')
        print('3. 应用翻译   (翻译表 -> 本地)')
        print('4. 撤回       (回退到上一次操作前)')
        print('0. 退出')
        choice = input('请选择: ').strip()
        if choice == '0':
            break
        elif choice == '1':
            action_add()
        elif choice == '2':
            action_import()
        elif choice == '3':
            run_mode(3, backup_objects=True)
        elif choice == '4':
            action_undo()
        else:
            print('无效选择')


# ---------- main ----------

def main():
    dry_run = '--dry-run' in sys.argv
    ignore_errors = '--ignore-errors' in sys.argv
    mode = None
    for a in sys.argv[1:]:
        if a.startswith('--mode='):
            mode = int(a.split('=', 1)[1])
        elif a == '--mode':
            pass
    # --mode N 形式
    if mode is None:
        args = [a for a in sys.argv[1:] if not a.startswith('--')]
        for a in args:
            if a in ('1', '2', '3'):
                mode = int(a)
                break
    if mode is None:
        sync_menu()
        return
    run_mode(mode, ignore_errors=ignore_errors, dry_run=dry_run)


if __name__ == '__main__':
    main()
